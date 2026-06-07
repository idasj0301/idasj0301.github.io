from __future__ import annotations

import json
import re
import shutil
from hashlib import sha256
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from html import escape, unescape
from pathlib import Path
from urllib.parse import quote, urljoin, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Volumes/Ida的硬盘，丢了赔付/单船票小程序上架-致ida.xlsx")
PRODUCTS_PATH = ROOT / "data" / "products.json"
DETAILS_DIR = ROOT / "data" / "details"
TRIPS_DIR = ROOT / "public" / "trips"
SHIPS_PATH = ROOT / "data" / "ships.json"
SHIP_PROFILES_DIR = ROOT / "data" / "ship-profiles"
OFFICIAL_IMAGES_PATH = ROOT / "data" / "ticket-official-images.json"
LOCAL_OFFICIAL_IMAGES_PATH = ROOT / "data" / "ticket-official-local-images.json"
LOCAL_OFFICIAL_IMAGES_DIR = ROOT / "public" / "ticket-official-images"
SOURCE_LABEL = "单船票小程序上架-致ida.xlsx"


SEABOURN_CARD_IMAGE_POOL = [
    "https://www.seabourn.com/content/dam/sbn65/inventory-assets/destinations/a/antarctica/antarctica-zodiac-landing_1066x800.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/destinations/A/antarctica/SBN_cardcarousel23-antarctica_042125.jpg",
    "https://www.seabourn.com/content/dam/sbn/itinerary/s6j20aan3/p621/s6j20aan3-p621.jpg",
    "https://www.seabourn.com/content/dam/sbn65/inventory-assets/ships/expedition-ships/venture/sbn-venture-kayak-teaser-newworlds-021926.jpg",
    "https://www.seabourn.com/content/dam/sbn/marketing-assets/landing-pages/pursuit-launch/r_Spotlight_Pursuit_Zodiacs_073123.jpg",
    "https://www.seabourn.com/content/dam/sbn/marketing-assets/landing-pages/pursuit-launch/r_Spotlight_Pursuit_Adventure_073123.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/itineraries/E2R28BAR7/V251A/arctic-ice-1334x1001.jpg",
    "https://www.seabourn.com/content/sbn/global/en/cruise-ships/seabourn-venture/2/_jcr_content/root/container/hero/image.coreimg.jpeg/1664833357827/vn-hero-2880x1047-v1-221003.jpeg",
    "https://www.seabourn.com/content/dam/sbn/marketing-assets/videos/thumbnails/c040-thumbs/video-venture-740x555.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/ships/cdp-banners/SBN_ship-slider_seabourn-pursuit-smoke-ceremony_090224.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/ships/cdp-banners/SBN_ship-slider_seabourn-pursuit_090224.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/ships/cdp-banners/SBN_ship-slider_expedition-ship-pool-ocean_090224.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/ships/cdp-banners/SBN_ship-slider_expedition-seabourn-pursuit-sauna-south-pacific_092224.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/activity-types/activity/the-club/SBN_2880x1047_hero_Club_PSVN_043025.jpg",
    "https://www.seabourn.com/content/dam/sbn65/inventory-assets/onboard-activities/activity/expedition-discovery-center/discovery-center_hero-v2_2560x575_310523.jpg",
    "https://www.seabourn.com/content/dam/sbn/inventory-assets/activity-types/bars-lounges/sky-bar/SBN_2880x1047_hero_SkyBar_PSVN_043025.jpg",
]


COMPANIES = {
    "夸克": {
        "id": "quark",
        "prefix": "quark",
        "company": "夸克探险 Quark Expeditions",
        "name": "夸克探险",
        "nameEn": "Quark Expeditions",
        "currency": "USD",
        "tags": ["探险", "南极", "摄影"],
        "logo": "/ships/quark.svg",
        "match": ["夸克", "Quark", "World Voyager", "Ocean Explorer", "Ultramarine"],
        "blurb": "老牌极地探险船司，覆盖南极半岛、南极飞航、雪丘岛等深度航线。",
    },
    "庞洛": {
        "id": "ponant",
        "prefix": "ponant",
        "company": "庞洛邮轮 PONANT",
        "name": "庞洛邮轮",
        "nameEn": "PONANT",
        "currency": "EUR",
        "tags": ["法式", "奢华", "南极"],
        "logo": "/ships/ponant.jpg",
        "match": ["庞洛", "PONANT", "Le Boréal", "L'Austral", "Le Lyrial"],
        "blurb": "法式小型奢华探险船队，提供南极半岛、南极圈、南乔治亚等航线。",
    },
    "ATLAS": {
        "id": "atlas",
        "prefix": "atlas",
        "company": "ATLAS 海洋邮轮 Atlas Ocean Voyages",
        "name": "ATLAS 海洋邮轮",
        "nameEn": "Atlas Ocean Voyages",
        "currency": "USD",
        "tags": ["探险", "南极", "小型船"],
        "logo": "/ships/atlas.png",
        "match": ["Atlas", "World Navigator", "World Traveller", "世界航海家", "世界旅行者"],
        "blurb": "小型奢华探险船队，南极半岛、南极圈及南乔治亚航次选择丰富。",
    },
    "A21": {
        "id": "antarctica21",
        "prefix": "a21",
        "company": "南极21 Antarctica21",
        "name": "南极21",
        "nameEn": "Antarctica21",
        "currency": "USD",
        "tags": ["飞航", "南极", "小型船"],
        "logo": "/ships/antarctica21.webp",
        "logoSource": "/Users/pro/Documents/官网/船司2/antarctica21.webp",
        "match": ["Antarctica21", "A21", "Magellan Explorer", "Magellan Discoverer", "麦哲伦"],
        "blurb": "南极飞航代表船司，以麦哲伦探索号、麦哲伦发现号服务半岛双飞和穿圈飞航。",
    },
    "66度": {
        "id": "seaventure",
        "prefix": "66exp",
        "company": "66度探险 66° Expeditions",
        "name": "66度探险 · 海神号",
        "nameEn": "66° Expeditions · Seaventure",
        "currency": "USD",
        "tags": ["中文服务", "南极", "海神号"],
        "logo": "/ships/expeditions-66.png",
        "match": ["66度", "66°", "SEAVENTURE", "Seaventure", "海神号"],
        "blurb": "以 Seaventure 海神号运营南极半岛、南极圈、南极双岛与三岛航线。",
    },
    "世邦": {
        "id": "seabourn",
        "prefix": "seabourn",
        "company": "世邦邮轮 Seabourn",
        "name": "世邦邮轮",
        "nameEn": "Seabourn",
        "currency": "USD",
        "tags": ["奢华", "探险", "南极"],
        "logo": "/ships/seabourn.png",
        "logoSource": "/Users/pro/Documents/官网/船司2/seabourn.png",
        "match": ["世邦", "Seabourn", "Venture", "Pursuit"],
        "blurb": "超奢华探险船队，世邦探险号与世邦追寻号覆盖南极半岛、双岛和三岛。",
    },
    "swan": {
        "id": "swan_hellenic",
        "prefix": "swan",
        "company": "天鹅海伦 Swan Hellenic",
        "name": "天鹅海伦",
        "nameEn": "Swan Hellenic",
        "currency": "USD",
        "tags": ["探险", "精品", "南极"],
        "logo": "/ships/swan_hellenic.png",
        "logoSource": "/Users/pro/Documents/官网/船司2/swan_hellenic.png",
        "match": ["Swan", "Swan Hellenic", "SH Diana", "SH Vega", "SH Minerva", "天鹅海伦"],
        "blurb": "精品探险船队，以 SH Diana、SH Vega、SH Minerva 运营南极半岛与威德尔海航线。",
    },
    "欧若拉": {
        "id": "aurora",
        "prefix": "aurora",
        "company": "欧若拉探险 Aurora Expeditions",
        "name": "欧若拉探险",
        "nameEn": "Aurora Expeditions",
        "currency": "USD",
        "tags": ["澳洲", "探险", "南极"],
        "logo": "/ships/aurora.svg",
        "logoSource": "/Users/pro/Documents/官网/船司2/aurora.svg",
        "match": ["欧若拉", "Aurora", "Douglas Mawson", "Sylvia Earle", "Greg Mortimer"],
        "blurb": "澳洲探险船司，覆盖经典南极半岛、南极圈、南极飞航、双岛与三岛航线。",
    },
}


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_ship(value: str) -> str:
    value = clean_text(value)
    value = value.replace("SEAVENTURE", "Seaventure")
    value = value.replace("SH DIANA", "SH Diana").replace("SH VEGA", "SH Vega")
    value = value.replace("  ", " ")
    value = re.sub(r"([A-Za-z][A-Za-z .'\-]*?)([\u4e00-\u9fff])", r"\1 \2", value)
    value = re.sub(r"([\u4e00-\u9fff])([A-Za-z])", r"\1 \2", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_route(value: str) -> str:
    value = clean_text(value)
    value = re.sub(r"\s*/\s*", " / ", value)
    return value.strip()


def to_iso_date(value, year_hint: int | None) -> str:
    if isinstance(value, datetime):
        dt = value
    else:
        text = clean_text(value)
        if not text:
            raise ValueError("missing date")
        dt = datetime.fromisoformat(text[:10])
    if year_hint and dt.year != year_hint:
        dt = dt.replace(year=year_hint)
    return dt.date().isoformat()


def duration_from_route(route: str) -> int:
    if match := re.search(r"(\d+)\s*晚", route):
        return int(match.group(1)) + 1
    if match := re.search(r"(\d+)\s*天", route):
        return int(match.group(1))
    if "双飞南极半岛" in route:
        return 8
    if "单飞南极半岛" in route:
        return 10
    if "雪丘岛" in route:
        return 14
    return 10


def end_date(start_iso: str, duration: int) -> str:
    start = datetime.fromisoformat(start_iso)
    return (start + timedelta(days=duration - 1)).date().isoformat()


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"https?://", "", text)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:70].strip("-") or "ticket"


def link_key(url: str) -> str:
    if "product_id=" in url:
        match = re.search(r"product_id=(\d+)", url)
        if match:
            return f"product-{match.group(1)}"
    if "voyageId=" in url:
        match = re.search(r"voyageId=(\d+)", url)
        if match:
            return f"voyage-{match.group(1)}"
    tail = url.split("?")[0].rstrip("/").split("/")[-1]
    return slugify(tail)


def canonical_url(url: str) -> str:
    if "product_id=" in url:
        match = re.search(r"product_id=(\d+)", url)
        if match:
            return f"{url.split('?')[0]}?product_id={match.group(1)}"
    if "voyageId=" in url:
        match = re.search(r"voyageId=(\d+)", url)
        if match:
            return f"{url.split('?')[0]}?voyageId={match.group(1)}"
    return url.split("?")[0]


def image_candidate_score(url: str) -> int:
    lower = url.lower()
    if any(char in url for char in ('"', "'", "\\", "{", "}", "<", ">")):
        return -100
    if not re.search(r"\.(?:jpe?g|png|webp)(?:[?#]|$)", lower):
        return -100
    score = 0
    for bad in ("logo", "favicon", "sprite", "icon", "badge", "placeholder", "transparent"):
        if bad in lower:
            score -= 40
    for good in (
        "antarctic",
        "antarctica",
        "expedition",
        "voyage",
        "cruise",
        "ship",
        "ice",
        "penguin",
        "glacier",
        "zodiac",
        "wildlife",
        "hero",
        "banner",
        "uploads",
        "content",
    ):
        if good in lower:
            score += 8
    if re.search(r"(?:1200|1600|1920|2048|2560|3000|1024)x", lower):
        score += 8
    return score


def clean_image_url(value: str, base_url: str) -> str:
    value = unescape(value or "").strip()
    if not value:
        return ""
    if any(char in value for char in ('"', "'", "\\", "{", "}", "<", ">")):
        return ""
    if "," in value and " " in value:
        value = value.split(",", 1)[0].strip().split(" ", 1)[0]
    return urljoin(base_url, value)


def normalize_official_image_url(url: str) -> str:
    parts = urlsplit(url)
    if parts.netloc == "res.cloudinary.com" and "/quark-web/" in parts.path:
        source_path = parts.path.split("/quark-web/", 1)[1]
        return f"https://assets.quarkexpeditions.com/uploads/{source_path}"
    return url


def extract_meta_image(markup: str, base_url: str) -> str:
    meta_candidates: list[str] = []
    for tag in re.findall(r"<meta\b[^>]*>", markup, flags=re.I):
        attrs = {
            key.lower(): value
            for key, value in re.findall(r"""([:\w-]+)\s*=\s*["']([^"']*)["']""", tag)
        }
        kind = (attrs.get("property") or attrs.get("name") or "").lower()
        content = attrs.get("content", "")
        if kind in {"og:image", "og:image:secure_url", "twitter:image", "twitter:image:src"}:
            candidate = clean_image_url(content, base_url)
            if candidate:
                meta_candidates.append(candidate)
    usable_meta = [url for url in meta_candidates if image_candidate_score(url) >= -20]
    if usable_meta:
        return max(usable_meta, key=image_candidate_score)

    candidates = [
        clean_image_url(match, base_url)
        for match in re.findall(
            r"""https?://[^"'<>\s\\,{}]+?\.(?:jpe?g|png|webp)(?:\?[^"'<>\s\\,{}]*)?""",
            markup,
            flags=re.I,
        )
    ]
    candidates.extend(
        clean_image_url(match, base_url)
        for match in re.findall(r"""(?:src|data-src|data-image|poster)\s*=\s*["']([^"']+\.(?:jpe?g|png|webp)(?:\?[^"']*)?)["']""", markup, flags=re.I)
    )
    candidates.extend(
        clean_image_url(match, base_url)
        for match in re.findall(
            r'''"(?:image|imageUrl|heroImage|primaryImage|linkImage|bannerImage|cardImage|thumbnail)"\s*:\s*"([^"]+\.(?:jpe?g|png|webp)(?:\?[^"]*)?)"''',
            markup,
            flags=re.I,
        )
    )
    candidates = list(dict.fromkeys(candidate for candidate in candidates if candidate))
    usable = [url for url in candidates if image_candidate_score(url) >= 0]
    return max(usable, key=image_candidate_score) if usable else ""


def fetch_antarctica21_image(url: str) -> str:
    match = re.search(r"product_id=(\d+)", url)
    if not match:
        return ""
    api_url = f"https://xdtz-oh4i-3nzx.n7d.xano.io/api:XeSy1hTp/cabin/{match.group(1)}"
    request = Request(api_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=8) as response:
        payload = json.loads(response.read().decode("utf-8", errors="ignore"))
    for item in payload if isinstance(payload, list) else []:
        content = item.get("_product_content") or {}
        for key in ("wf_product_image", "wf_product_itinerary_image"):
            image = content.get(key, "")
            if image:
                return image
    return ""


def fetch_66expeditions_image(url: str) -> str:
    match = re.search(r"voyageId=(\d+)", url)
    if not match:
        return ""
    body = json.dumps({"langCode": "zh-CN", "voyageId": match.group(1)}).encode("utf-8")
    request = Request(
        "https://www.66expeditions.com/prod-api/web/api/ofWeb/routesDetail",
        data=body,
        method="POST",
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
    )
    with urlopen(request, timeout=8) as response:
        payload = json.loads(response.read().decode("utf-8", errors="ignore"))
    data = payload.get("data") or {}
    for key in ("bannerPicUrl", "picUrl", "routePicUrl"):
        value = data.get(key, "")
        if value:
            return value.split(",", 1)[0].strip()
    for item in data.get("logList") or []:
        value = item.get("imageUrl", "")
        if value:
            return value.split(",", 1)[0].strip()
    return ""


def fetch_domain_specific_image(url: str) -> str:
    if "antarctica21.com/booking/quote-request" in url:
        return fetch_antarctica21_image(url)
    if "66expeditions.com/routeDetails" in url:
        return fetch_66expeditions_image(url)
    return ""


def load_official_image_cache() -> dict[str, str]:
    if not OFFICIAL_IMAGES_PATH.exists():
        return {}
    return json.loads(OFFICIAL_IMAGES_PATH.read_text())


def load_local_official_image_cache() -> dict[str, str]:
    if not LOCAL_OFFICIAL_IMAGES_PATH.exists():
        return {}
    return json.loads(LOCAL_OFFICIAL_IMAGES_PATH.read_text())


def fetch_official_image(url: str, cache: dict[str, str]) -> str:
    if cache.get(url):
        cache[url] = normalize_official_image_url(cache[url])
        return cache[url]
    try:
        image = fetch_domain_specific_image(url)
        if not image:
            request = Request(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
                },
            )
            with urlopen(request, timeout=8) as response:
                raw = response.read(3_000_000)
                charset = response.headers.get_content_charset() or "utf-8"
            image = extract_meta_image(raw.decode(charset, errors="ignore"), url)
    except Exception as exc:  # noqa: BLE001 - keep syncing other SKUs even when one vendor blocks.
        print(f"Warning: official image fetch failed for {url}: {exc}")
        image = ""
    cache[url] = normalize_official_image_url(image)
    return cache[url]


def requestable_url(url: str) -> str:
    parts = urlsplit(url)
    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            quote(parts.path, safe="/%"),
            quote(parts.query, safe="=&%?/:,+"),
            quote(parts.fragment, safe="=&%?/:,+"),
        )
    )


def image_extension(url: str, content_type: str, data: bytes) -> str:
    lower_type = content_type.lower()
    if "webp" in lower_type or data.startswith(b"RIFF") and b"WEBP" in data[:16]:
        return "webp"
    if "png" in lower_type or data.startswith(b"\x89PNG"):
        return "png"
    if "jpeg" in lower_type or "jpg" in lower_type or data.startswith(b"\xff\xd8"):
        return "jpg"
    path = urlsplit(url).path.lower()
    if match := re.search(r"\.(jpe?g|png|webp)$", path):
        ext = match.group(1)
        return "jpg" if ext == "jpeg" else ext
    return "jpg"


def download_official_image(url: str, local_cache: dict[str, str]) -> str:
    cached = local_cache.get(url)
    if cached and (ROOT / "public" / cached.lstrip("/")).exists():
        return cached
    request = Request(
        requestable_url(url),
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        },
    )
    with urlopen(request, timeout=20) as response:
        data = response.read(8_000_000)
        content_type = response.headers.get("content-type", "")
    if len(data) < 512:
        raise ValueError("downloaded image is unexpectedly small")
    if not (
        "image/" in content_type.lower()
        or data.startswith(b"\xff\xd8")
        or data.startswith(b"\x89PNG")
        or (data.startswith(b"RIFF") and b"WEBP" in data[:16])
    ):
        raise ValueError(f"downloaded file is not an image: {content_type or 'unknown content-type'}")
    ext = image_extension(url, content_type, data)
    digest = sha256(url.encode("utf-8")).hexdigest()[:18]
    LOCAL_OFFICIAL_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    target = LOCAL_OFFICIAL_IMAGES_DIR / f"{digest}.{ext}"
    target.write_bytes(data)
    local_path = f"/ticket-official-images/{target.name}"
    local_cache[url] = local_path
    return local_path


def attach_official_images(items: list[dict]) -> None:
    cache = load_official_image_cache()
    changed = False
    for url, image in list(cache.items()):
        normalized = normalize_official_image_url(image)
        if normalized != image:
            cache[url] = normalized
            changed = True
        if normalized and image_candidate_score(normalized) < -20:
            cache[url] = ""
            changed = True
    if changed:
        OFFICIAL_IMAGES_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    urls = sorted({item["url"] for item in items})
    missing = [url for url in urls if not cache.get(url)]
    if missing:
        print(f"Fetching official images: {len(missing)} uncached links")
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(fetch_official_image, url, cache): url for url in missing}
            for index, future in enumerate(as_completed(futures), start=1):
                url = futures[future]
                image = future.result()
                status = "ok" if image else "miss"
                print(f"[{index}/{len(missing)}] {status}: {url}")
                OFFICIAL_IMAGES_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    elif not OFFICIAL_IMAGES_PATH.exists():
        OFFICIAL_IMAGES_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    normalized_cache = {url: normalize_official_image_url(image) for url, image in cache.items()}
    if normalized_cache != cache:
        cache.clear()
        cache.update(normalized_cache)
        OFFICIAL_IMAGES_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    for item in items:
        item["officialImage"] = cache.get(item["url"], "")


def diversify_seabourn_card_images(items: list[dict]) -> None:
    seabourn_items = sorted(
        (item for item in items if item.get("companyId") == "seabourn"),
        key=lambda item: (item["departureDate"], item["slug"]),
    )
    if not seabourn_items:
        return
    cache = load_official_image_cache()
    for index, item in enumerate(seabourn_items):
        image = SEABOURN_CARD_IMAGE_POOL[index % len(SEABOURN_CARD_IMAGE_POOL)]
        item["officialImage"] = image
        cache[item["url"]] = image
    OFFICIAL_IMAGES_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")


def attach_local_official_images(items: list[dict]) -> None:
    local_cache = load_local_official_image_cache()
    official_urls = sorted({item["officialImage"] for item in items if item.get("officialImage")})
    missing = [
        url
        for url in official_urls
        if not local_cache.get(url) or not (ROOT / "public" / local_cache[url].lstrip("/")).exists()
    ]
    if missing:
        print(f"Downloading official images locally: {len(missing)} uncached images")
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(download_official_image, url, local_cache): url for url in missing}
            for index, future in enumerate(as_completed(futures), start=1):
                url = futures[future]
                try:
                    local_path = future.result()
                    print(f"[{index}/{len(missing)}] cached: {local_path}")
                except Exception as exc:  # noqa: BLE001 - preserve remote image fallback if caching fails.
                    print(f"Warning: local image cache failed for {url}: {exc}")
                LOCAL_OFFICIAL_IMAGES_PATH.write_text(
                    json.dumps(local_cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
                )
    elif not LOCAL_OFFICIAL_IMAGES_PATH.exists():
        LOCAL_OFFICIAL_IMAGES_PATH.write_text(json.dumps(local_cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    for item in items:
        official_image = item.get("officialImage", "")
        item["localOfficialImage"] = local_cache.get(official_image, official_image)


def price_value(value) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    if not text or "售罄" in text:
        return 0
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0


def price_label(price: float, currency: str) -> str:
    if price <= 0:
        return "价格咨询"
    amount = f"{price:,.0f}" if price == int(price) else f"{price:,.2f}"
    if currency == "EUR":
        return f"€{amount} 起"
    return f"US${amount} 起"


def route_destination(route: str) -> str:
    if "三岛" in route or "南极3岛" in route or "南极三岛" in route:
        return "福克兰群岛 / 南乔治亚 / 南极半岛"
    if "双岛" in route or "南乔治亚" in route:
        return "南乔治亚 / 南极半岛"
    if "雪丘" in route:
        return "雪丘岛 / 南极半岛"
    if "威德尔" in route:
        return "威德尔海 / 南极半岛"
    if "穿圈" in route or "南极圈" in route:
        return "南极圈 / 南极半岛"
    if "飞" in route:
        return "南极半岛 / 南设得兰群岛 / 飞航衔接"
    return "南极半岛"


def route_kind(route: str) -> str:
    if "三岛" in route or "3岛" in route:
        return "three-islands"
    if "双岛" in route or "南乔治亚" in route:
        return "south-georgia"
    if "雪丘" in route:
        return "snow-hill"
    if "威德尔" in route:
        return "weddell"
    if "穿圈" in route or "南极圈" in route:
        return "circle"
    if "飞" in route:
        return "fly"
    return "peninsula"


def compact_itinerary(route: str, start_iso: str, duration: int) -> list[dict]:
    kind = route_kind(route)
    start_label = cn_short_date(start_iso)
    finish_label = cn_short_date(end_date(start_iso, duration))
    if kind == "fly":
        return [
            {"day": "1", "date": start_label, "title": "飞航衔接与登船", "content": "按船司飞航或接驳安排抵达登船点，开启南极单船票航段。"},
            {"day": "2", "title": "南极半岛海域", "content": "进入南极半岛或南设得兰群岛海域，参加船司安排的巡游、登陆或讲座。"},
            {"day": str(max(duration - 2, 3)), "title": "南极探索日", "content": "在南极海峡、半岛、岛屿或冰山水域探索，具体登陆点以天气、冰况和船长安排为准。"},
            {"day": str(duration), "date": finish_label, "title": "离船", "content": "按船司安排离船或衔接返程飞航，结束邮轮航段。"},
        ]
    if kind in {"three-islands", "south-georgia"}:
        return [
            {"day": "1", "date": start_label, "title": "登船", "content": "按船司指定港口登船，开启南大西洋与南极单船票航段。"},
            {"day": "4", "title": "海上航行与讲座", "content": "海上航行日通常安排安全说明、目的地讲座和野生动物观察。"},
            {"day": str(max(duration // 2, 6)), "title": "南乔治亚 / 岛屿探索", "content": "探索南乔治亚或相关岛屿海域，重点关注企鹅、海豹、海鸟与探险史遗迹。"},
            {"day": str(max(duration - 4, 8)), "title": "南极半岛海域", "content": "进入南极半岛、南设得兰群岛或南极海峡，参加巡游、登陆或观景活动。"},
            {"day": str(duration), "date": finish_label, "title": "离船", "content": "抵达离船港并结束邮轮航段。"},
        ]
    middle_title = {
        "snow-hill": "雪丘岛方向探索",
        "weddell": "威德尔海方向探索",
        "circle": "南极圈与半岛探索",
    }.get(kind, "南极半岛探索")
    return [
        {"day": "1", "date": start_label, "title": "登船", "content": "按船司指定港口登船，开启南极单船票航段。"},
        {"day": "3", "title": "德雷克海峡 / 海上航行", "content": "海上航行日通常安排探险队讲座、登陆说明和观景活动。"},
        {"day": str(max(duration // 2, 4)), "title": middle_title, "content": "在南极半岛、岛屿、冰山水域或相关主题海域探索，活动以天气和冰况为准。"},
        {"day": str(duration), "date": finish_label, "title": "离船", "content": "抵达离船港并结束邮轮航段。"},
    ]


def cn_short_date(iso: str) -> str:
    dt = datetime.fromisoformat(iso)
    return f"{dt.month}月{dt.day}日"


def date_range(start: str, end: str) -> str:
    return f"{start.replace('-', '.')} — {end.replace('-', '.')}"


def build_hero_svg(item: dict) -> str:
    title = escape(item["title"])
    company = escape(item["company"])
    route = escape(item["route"])
    ship = escape(item["shipName"])
    dates = escape(date_range(item["departureDate"], item["endDate"]))
    price = escape(item["priceLabel"])
    color = item["accent"]
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1600 900" role="img" aria-labelledby="title desc">
  <title id="title">{title}</title>
  <desc id="desc">{company} · {ship} · {dates}</desc>
  <defs>
    <linearGradient id="sky" x1="0" x2="1" y1="0" y2="1">
      <stop offset="0" stop-color="#f7fbff"/>
      <stop offset=".56" stop-color="#dcebf3"/>
      <stop offset="1" stop-color="#b7d0dc"/>
    </linearGradient>
    <linearGradient id="sea" x1="0" x2="0" y1="0" y2="1">
      <stop offset="0" stop-color="{color}"/>
      <stop offset="1" stop-color="#061d34"/>
    </linearGradient>
  </defs>
  <rect width="1600" height="900" fill="url(#sky)"/>
  <path d="M0 585 C170 520 300 610 455 548 C610 486 775 622 950 560 C1150 490 1320 610 1600 515 L1600 900 L0 900 Z" fill="url(#sea)"/>
  <path d="M160 530 L360 340 L500 505 L630 300 L785 522 Z" fill="#ffffff" opacity=".86"/>
  <path d="M1030 560 L1178 360 L1260 505 L1370 330 L1505 570 Z" fill="#ffffff" opacity=".72"/>
  <circle cx="1290" cy="170" r="78" fill="#ffffff" opacity=".65"/>
  <g transform="translate(96 95)">
    <text x="0" y="0" class="kicker">{company}</text>
    <text x="0" y="88" class="title">{title}</text>
    <text x="0" y="150" class="sub">{ship}</text>
    <text x="0" y="205" class="meta">{route} · {dates}</text>
  </g>
  <g transform="translate(96 706)">
    <rect width="410" height="82" rx="41" fill="#ffffff" opacity=".92"/>
    <text x="34" y="52" class="price">{price}</text>
  </g>
  <style>
    .kicker {{ font: 700 30px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: {color}; }}
    .title {{ font: 760 72px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #101820; }}
    .sub {{ font: 500 34px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #46525d; }}
    .meta {{ font: 500 30px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #5f6d78; }}
    .price {{ font: 760 32px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #101820; }}
  </style>
</svg>
"""


def build_route_svg(item: dict) -> str:
    points_by_kind = {
        "fly": [(610, 135), (560, 245), (470, 380), (630, 275), (610, 135)],
        "three-islands": [(755, 515), (790, 350), (590, 365), (545, 300), (470, 380), (630, 275), (755, 515)],
        "south-georgia": [(755, 515), (590, 365), (545, 300), (470, 380), (755, 515)],
        "snow-hill": [(755, 515), (650, 430), (590, 300), (520, 245), (470, 380), (755, 515)],
        "weddell": [(755, 515), (650, 430), (560, 245), (520, 190), (470, 380), (755, 515)],
        "circle": [(755, 515), (650, 430), (470, 380), (385, 440), (630, 275), (755, 515)],
        "peninsula": [(755, 515), (650, 430), (560, 245), (470, 380), (630, 275), (755, 515)],
    }
    labels = [
        ("登船 / 离船港", 755, 515),
        ("南极海峡", 560, 245),
        ("南极半岛", 470, 380),
        ("南设得兰群岛", 630, 275),
        ("南乔治亚", 590, 365),
        ("南极圈方向", 385, 440),
        ("雪丘岛 / 威德尔海方向", 520, 190),
    ]
    kind = route_kind(item["route"])
    points = points_by_kind.get(kind, points_by_kind["peninsula"])
    path = " ".join(f"{'M' if idx == 0 else 'L'} {x} {y}" for idx, (x, y) in enumerate(points))
    label_nodes = "\n".join(
        f'<g><circle cx="{x}" cy="{y}" r="9" fill="{item["accent"]}" stroke="#fff" stroke-width="4"/><text x="{x + 15}" y="{y - 12}" class="label">{escape(label)}</text></g>'
        for label, x, y in labels
    )
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 960 620" role="img" aria-labelledby="title desc">
  <title id="title">{escape(item["title"])} 航线图</title>
  <desc id="desc">{escape(item["route"])} · {escape(date_range(item["departureDate"], item["endDate"]))}</desc>
  <rect width="960" height="620" rx="28" fill="#edf6fb"/>
  <path d="M0 465 C140 418 228 475 340 430 C470 378 548 505 670 470 C805 428 842 516 960 476 L960 620 L0 620 Z" fill="#d8e4e8"/>
  <path d="{path}" fill="none" stroke="{item["accent"]}" stroke-width="7" stroke-linecap="round" stroke-linejoin="round" stroke-dasharray="1 15"/>
  <path d="{path}" fill="none" stroke="#c8a45d" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>
  {label_nodes}
  <g transform="translate(54 58)">
    <rect width="560" height="122" rx="18" fill="#ffffff" opacity=".94"/>
    <text x="24" y="36" class="kicker">{escape(item["company"])}</text>
    <text x="24" y="70" class="title">{escape(item["route"])}</text>
    <text x="24" y="98" class="small">{escape(item["shipName"])} · {escape(date_range(item["departureDate"], item["endDate"]))}</text>
  </g>
  <text x="54" y="572" class="small">航线为单船票咨询示意；实际港口、登陆点、飞航/接驳以船司最终确认为准。</text>
  <style>
    .title {{ font: 700 23px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #102233; }}
    .kicker {{ font: 700 13px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; letter-spacing: .08em; fill: {item["accent"]}; }}
    .label {{ font: 600 16px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #0f2638; paint-order: stroke; stroke: #fff; stroke-width: 5px; stroke-linejoin: round; }}
    .small {{ font: 15px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #415668; }}
  </style>
</svg>
"""


def build_gallery_svg(item: dict, index: int) -> str:
    blocks = [
        ("航线", item["route"], route_destination(item["route"])),
        ("邮轮", item["shipName"], item["company"]),
        ("单船票", item["priceLabel"], "舱位、税费和余舱以顾问实时确认为准"),
    ]
    kicker, title, sub = blocks[index]
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1200 800">
  <rect width="1200" height="800" fill="#f5f7fa"/>
  <circle cx="980" cy="165" r="120" fill="{item["accent"]}" opacity=".13"/>
  <path d="M0 560 C180 500 320 610 500 545 C710 470 850 610 1200 500 L1200 800 L0 800 Z" fill="{item["accent"]}" opacity=".9"/>
  <text x="78" y="120" class="kicker">{escape(kicker)}</text>
  <text x="78" y="215" class="title">{escape(title)}</text>
  <text x="78" y="282" class="sub">{escape(sub)}</text>
  <style>
    .kicker {{ font: 700 30px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: {item["accent"]}; }}
    .title {{ font: 760 64px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #15191f; }}
    .sub {{ font: 500 34px -apple-system, BlinkMacSystemFont, "PingFang SC", sans-serif; fill: #66717b; }}
  </style>
</svg>
"""


def product_for(item: dict) -> dict:
    route = item["route"]
    summary = (
        f"{item['title']}，{cn_full_date(item['departureDate'])}至{cn_full_date(item['endDate'])}，"
        f"{item['durationDays']}天，搭乘{item['shipName']}。"
        f"蘑菇表格参考起价 {item['priceLabel']}；官方链接已记录，最终舱位、价格、税费和包含项以船司实时确认为准。"
    )
    return {
        "id": item["id"],
        "slug": item["slug"],
        "title": item["title"],
        "category": "ticket",
        "subcategory": route,
        "tags": [
            item["company"],
            "单船票",
            "南极",
            item["companyEn"],
            item["shipName"],
            item["departureDate"],
        ],
        "departureDate": item["departureDate"],
        "endDate": item["endDate"],
        "durationDays": item["durationDays"],
        "priceFrom": item["priceFrom"],
        "priceLabel": item["priceLabel"],
        "company": item["company"],
        "shipName": item["shipName"],
        "summary": summary,
        "overview": summary,
        "highlights": [
            f"{route} · {route_destination(route)}",
            f"{item['company']} · {item['shipName']}",
            f"蘑菇表格参考起价：{item['priceLabel']}",
            "本页为单船票咨询 SKU，舱位和价格需向船司实时确认。",
        ],
        "itinerary": [
            {"day": int(day["day"]), "title": day["title"], "content": day["content"]}
            for day in compact_itinerary(route, item["departureDate"], item["durationDays"])
            if day["day"].isdigit()
        ],
        "ship": f"{item['shipName']}：所属船司 {item['company']}，具体舱房、餐饮、公共区域和探险设施以船司官方资料及签约合同为准。",
        "feeNote": "单船票价格、舱位、税费、接驳/飞航衔接和附加服务以顾问向船司实时确认为准；国际机票、签证、保险及个人消费通常另计。",
        "notice": "南极航线、飞航、登陆和巡游活动受天气、冰况、港口、机场与船司安排影响，航线、登陆点和活动顺序可能调整。",
        "published": True,
        "featured": False,
        "imageAlt": item["title"],
        "sourceFile": f"{SOURCE_LABEL} / {item['url']}",
        "wecomFrom": item["id"],
    }


def detail_for(item: dict) -> dict:
    route = item["route"]
    itinerary = compact_itinerary(route, item["departureDate"], item["durationDays"])
    official_image = item.get("localOfficialImage") or item.get("officialImage") or f"/trips/{item['slug']}/hero.svg"
    return {
        "titleEn": item["url"].split("?")[0].rstrip("/").split("/")[-1].replace("-", " ").title(),
        "subtitle": route,
        "tags": [item["company"], "单船票", "南极", item["shipName"], item["departureDate"]],
        "cardImage": official_image,
        "heroImage": official_image,
        "routeMap": {
            "src": f"/trips/{item['slug']}/route-map.svg",
            "alt": f"{item['title']} 航线图",
            "caption": f"{route} · {date_range(item['departureDate'], item['endDate'])}",
        },
        "gallery": [
            {"src": official_image, "alt": f"{item['title']} 船司官网产品图", "caption": "船司官网产品图"},
            {"src": f"/trips/{item['slug']}/gallery-01.svg", "alt": f"{route} 航线说明", "caption": "航线主题"},
            {"src": f"/trips/{item['slug']}/gallery-02.svg", "alt": f"{item['shipName']} 邮轮说明", "caption": "邮轮与船司"},
        ],
        "metaTable": [
            ["船司", item["company"]],
            ["邮轮", item["shipName"]],
            ["航线", route],
            ["目的地", route_destination(route)],
            ["出行日期", date_range(item["departureDate"], item["endDate"])],
            ["出行时长", f"{item['durationDays']}天"],
            ["参考起价", item["priceLabel"]],
            ["航线来源", item["url"]],
        ],
        "cabins": [
            {"name": "基础舱型 / 标准舱型", "spec": "具体开放舱型、面积、楼层和权益以船司实时库存为准", "price": item["priceLabel"]},
            {"name": "阳台房 / 套房 / 高阶套房", "spec": "不同船、航次和促销政策价格差异较大，请顾问实时确认", "price": "实时确认"},
        ],
        "highlightSections": [
            {
                "title": "航线亮点",
                "content": f"{route} 单船票航程，目的地覆盖 {route_destination(route)}。本页根据蘑菇表格提供的官方链接、航次日期、邮轮和起价整理；具体每日港口和登陆安排以船司官方最终文件为准。",
            },
            {
                "title": "船司与邮轮",
                "bullets": [
                    f"船司：{item['company']}",
                    f"邮轮：{item['shipName']}",
                    "船上住宿、餐饮、探险队、讲座、Zodiac 巡游/登陆等项目以船司该航次开放内容为准。",
                ],
            },
            {
                "title": "单船票定位",
                "bullets": [
                    "适合已自行安排国际机票、签证、保险、前后住宿或岸上行程，只需要邮轮航段的客人。",
                    "官网展示为可咨询 SKU，不代表实时锁舱；舱位、税费、促销和取消条款需顾问向船司实时确认。",
                    "如涉及南极飞航，航班、接驳和前后缓冲时间需按船司最终安排预留。",
                ],
            },
            {
                "title": "价格与库存",
                "content": f"蘑菇表格参考起价为 {item['priceLabel']}。不同舱型、人数、汇率、税费和船司促销会影响最终报价，请以签约前顾问确认为准。",
            },
        ],
        "itinerary": itinerary,
        "itineraryNote": "以上为单船票咨询版行程框架。实际港口、每日时间、登陆点、冲锋艇巡游、徒步、飞航和接驳安排以船司官方最终文件、天气、冰况和船长/探险队长安排为准。",
        "shipDetail": {
            "name": item["shipName"],
            "intro": f"{item['shipName']} 服务于 {item['company']} 的 {route} 航次。船上设施、舱型和探险配置请以该船司官方页面、实时库存和签约合同为准。",
            "specs": [
                ["船司", item["company"]],
                ["邮轮", item["shipName"]],
                ["航线类型", route],
                ["出发日期", item["departureDate"]],
                ["来源链接", item["url"]],
            ],
            "facilities": [
                "船上住宿与餐饮服务",
                "探险队讲座、登陆说明与安全简报",
                "Zodiac 冲锋艇巡游/登陆活动（以航次开放为准）",
                "观景甲板、公共休息区及船司安排的船上活动",
            ],
            "cabins": [
                {"name": "开放舱型", "spec": "以船司实时库存、价格表和合同为准", "price": item["priceLabel"]},
            ],
        },
        "feeIncluded": [
            "船上住宿、船上餐食和船司安排的船上活动（以该航次条款为准）",
            "探险队讲座、登陆说明及船司开放的 Zodiac 巡游/登陆项目（以天气、冰况和船司安排为准）",
            "港务税费、船上小费、饮品、Wi-Fi 等是否包含，需按对应船司和舱型政策实时确认。",
        ],
        "feeExcluded": [
            "国际机票、签证、保险、出发前后酒店、接送机、岸上延伸行程及个人消费，除非船司或合同另有明确包含。",
            "皮划艇、露营、直升机、潜水等可选活动如需额外名额、条件或费用，以船司确认为准。",
            "因天气、冰况、航班、机场、港口或不可抗力造成的额外费用，按船司条款和合同约定执行。",
        ],
        "noticeSections": [
            {
                "title": "行程调整",
                "content": "极地航线高度依赖天气、冰况、港口、机场和环保规范。船长与探险队长会根据安全和体验调整航线、登陆点、时间和活动顺序。",
            },
            {
                "title": "单船票说明",
                "content": "本页用于官网咨询和顾问匹配；最终舱位、价格、税费、飞航/接驳和费用包含以船客顾问向船司实时确认为准。",
            },
        ],
    }


def cn_full_date(iso: str) -> str:
    dt = datetime.fromisoformat(iso)
    return f"{dt.year}年{dt.month}月{dt.day}日"


def read_rows() -> list[dict]:
    wb = load_workbook(WORKBOOK, data_only=True)
    rows = []
    accents = ["#003066", "#24516c", "#44606f", "#315f7a", "#4b6677", "#234b63", "#5a6d78", "#173d55"]
    for sheet_index, ws in enumerate(wb.worksheets):
        if ws.title == "银海双飞":
            continue
        company = COMPANIES.get(ws.title)
        if not company:
            continue
        route = ""
        year_hint = None
        for raw in ws.iter_rows(min_row=2, values_only=True):
            first = clean_text(raw[0] if len(raw) > 0 else "")
            if first == "其他":
                continue
            if clean_text(raw[1] if len(raw) > 1 else ""):
                route = normalize_route(raw[1])
            if clean_text(raw[2] if len(raw) > 2 else ""):
                try:
                    year_hint = int(float(clean_text(raw[2])))
                except ValueError:
                    year_hint = None
            if not route:
                continue
            ship = normalize_ship(raw[4] if len(raw) > 4 else "")
            link_col = 6 if ws.title == "庞洛" or ws.title == "ATLAS" else 5
            price_col = 5 if ws.title == "庞洛" or ws.title == "ATLAS" else 6
            url = clean_text(raw[link_col] if len(raw) > link_col else "")
            if not ship or not url.startswith("http"):
                continue
            start = to_iso_date(raw[3], year_hint)
            duration = duration_from_route(route)
            price = price_value(raw[price_col] if len(raw) > price_col else None)
            key = link_key(url)
            slug = f"{company['prefix']}-{start}-{slugify(ship)}-{key}"
            title = f"{ship} · {route}"
            rows.append(
                {
                    "id": f"tkt-{company['prefix']}-{len(rows) + 1:03d}",
                    "slug": slug,
                    "title": title,
                    "route": route,
                    "company": company["company"],
                    "companyEn": company["nameEn"],
                    "companyId": company["id"],
                    "shipName": ship,
                    "departureDate": start,
                    "endDate": end_date(start, duration),
                    "durationDays": duration,
                    "priceFrom": price,
                    "priceLabel": price_label(price, company["currency"]),
                    "url": canonical_url(url),
                    "accent": accents[sheet_index % len(accents)],
                }
            )
    return rows


def sync_products(items: list[dict]) -> None:
    generated_slugs = {item["slug"] for item in items}
    products = json.loads(PRODUCTS_PATH.read_text())
    kept = []
    for product in products:
        if product["slug"] in generated_slugs:
            continue
        if product.get("sourceFile", "").startswith(f"{SOURCE_LABEL} /"):
            continue
        if product.get("category") == "ticket" and product.get("company") == "银海":
            product["company"] = "银海 Silversea Cruises"
            product["tags"] = [
                "银海 Silversea Cruises" if tag == "银海" else tag
                for tag in product.get("tags", [])
            ]
        kept.append(product)
    new_products = [product_for(item) for item in items]
    insert_at = max((idx for idx, product in enumerate(kept) if product.get("category") == "ticket"), default=-1) + 1
    kept[insert_at:insert_at] = sorted(new_products, key=lambda p: (p["departureDate"], p["company"], p["shipName"]))
    PRODUCTS_PATH.write_text(json.dumps(kept, ensure_ascii=False, indent=2) + "\n")


def sync_details(items: list[dict]) -> None:
    DETAILS_DIR.mkdir(parents=True, exist_ok=True)
    current_slugs = {item["slug"] for item in items}
    prefixes = tuple(f"{company['prefix']}-" for company in COMPANIES.values())
    for file in DETAILS_DIR.glob("*.json"):
        slug = file.stem
        if slug.startswith(prefixes) and slug not in current_slugs:
            file.unlink()
    for trip_dir in TRIPS_DIR.iterdir():
        if trip_dir.is_dir() and trip_dir.name.startswith(prefixes) and trip_dir.name not in current_slugs:
            shutil.rmtree(trip_dir)
    for item in items:
        trip_dir = TRIPS_DIR / item["slug"]
        trip_dir.mkdir(parents=True, exist_ok=True)
        (trip_dir / "hero.svg").write_text(build_hero_svg(item))
        (trip_dir / "route-map.svg").write_text(build_route_svg(item))
        for idx in range(3):
            (trip_dir / f"gallery-{idx + 1:02d}.svg").write_text(build_gallery_svg(item, idx))
        (DETAILS_DIR / f"{item['slug']}.json").write_text(json.dumps(detail_for(item), ensure_ascii=False, indent=2) + "\n")


def sync_ship_logos_and_data(items: list[dict]) -> None:
    public_ships = ROOT / "public" / "ships"
    public_ships.mkdir(parents=True, exist_ok=True)
    for company in COMPANIES.values():
        source = company.get("logoSource")
        if source and Path(source).exists():
            shutil.copyfile(source, ROOT / "public" / company["logo"].lstrip("/"))

    ships = json.loads(SHIPS_PATH.read_text())
    by_id = {ship["id"]: ship for ship in ships}
    for company in COMPANIES.values():
        entry = {
            "id": company["id"],
            "name": company["name"],
            "nameEn": company["nameEn"],
            "logo": company["logo"],
            "tags": company["tags"],
            "match": company["match"],
            "blurb": company["blurb"],
        }
        if company["id"] in by_id:
            by_id[company["id"]].update(entry)
        else:
            ships.append(entry)
    SHIPS_PATH.write_text(json.dumps(ships, ensure_ascii=False, indent=2) + "\n")

    grouped: dict[str, list[dict]] = {}
    for item in items:
        grouped.setdefault(item["companyId"], []).append(item)
    SHIP_PROFILES_DIR.mkdir(parents=True, exist_ok=True)
    for company in COMPANIES.values():
        profile_path = SHIP_PROFILES_DIR / f"{company['id']}.json"
        if profile_path.exists():
            continue
        related = grouped.get(company["id"], [])[:8]
        vessels = sorted({item["shipName"] for item in grouped.get(company["id"], [])})
        profile = {
            "brandId": company["id"],
            "vesselName": f"{company['name']} · 单船票船队",
            "vesselNameEn": company["nameEn"],
            "tagline": company["blurb"],
            "heroImage": company["logo"],
            "highlights": [
                "官网单船票 SKU 已按蘑菇表格补充航线、日期、邮轮、起价和官方链接",
                "适合已自行安排国际段、签证和前后住宿，只需预订单船票航段的客人",
                "最终舱位、价格、税费、活动和取消条款以船司实时确认为准",
                "南极登陆与飞航安排受天气、冰况、港口、机场和环保规范影响",
            ],
            "sections": [
                {
                    "title": "船司定位",
                    "content": f"{company['name']}（{company['nameEn']}）是船客单船票可咨询船司之一。本页汇总其在售或可咨询的南极航次，方便客人按船司、船名、日期和预算进行筛选。",
                },
                {
                    "title": "咨询建议",
                    "content": "单船票适合明确出行月份、可自行安排国际段和前后住宿的客人。建议提交预算、舱型偏好、可出发日期和是否接受飞航/德雷克海峡航行，由顾问实时核对余舱。",
                },
            ],
            "specs": [
                ["船司", f"{company['name']} / {company['nameEn']}"],
                ["区域", "南极"],
                ["产品类型", "单船票"],
                ["库存", "以船司实时确认为准"],
            ],
            "facilities": [
                "船上住宿与餐饮",
                "探险队讲座与登陆说明",
                "Zodiac 冲锋艇巡游/登陆活动（以航次开放为准）",
                "观景甲板和公共休息空间",
            ],
            "vessels": [
                {
                    "name": vessel,
                    "role": "南极单船票航次",
                    "intro": f"{vessel} 出现在蘑菇表格的 {company['name']} 单船票 SKU 中，具体船舶参数、舱型和设施以船司官方资料为准。",
                }
                for vessel in vessels
            ],
            "relatedSlugs": [item["slug"] for item in related],
        }
        profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n")


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")
    items = read_rows()
    attach_official_images(items)
    diversify_seabourn_card_images(items)
    attach_local_official_images(items)
    sync_products(items)
    sync_details(items)
    sync_ship_logos_and_data(items)
    print(f"Synced {len(items)} non-Silversea ticket SKUs from {WORKBOOK}")


if __name__ == "__main__":
    main()
